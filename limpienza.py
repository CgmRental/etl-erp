from openpyxl import load_workbook
import sys
import os
import polars as pl

def main():

    formato = ".xlsx"
    ruta_archivos_base = r'C:\Users\CGM\Projects\Proyecto CGMExPress\data'
    nombre_archivo_entregas_ewm_cabezera = "MONITOR_CABEZERA"
    nombre_archivo_entregas_ewm_detalle = "MONITOR_DETALLE"
    nombre_archivo_guias_remision = "guia_remision"
    centros = pl.DataFrame({'centro': ['C200', 'C040', 'C080']})

    nombre_archivo_entregas_ewm_cabezera_pendiente = "MONITOR_CABEZERA_PENDIENTE"
    nombre_archivo_entregas_ewm_detalle_pendiente = "MONITOR_DETALLE_PENDIENTE"

    centros_pendiente = pl.DataFrame({'centro': ['C154','C200', 'C040', 'C080']})

    def limpiar_excel_openpyxl(ruta_archivo):
        wb = load_workbook(ruta_archivo)
        ws = wb.active
        
        # Borrar todas las filas excepto la primera
        ws.delete_rows(2, ws.max_row)
        
        wb.save(ruta_archivo)

    limpiar_excel_openpyxl(os.path.join(ruta_archivos_base,nombre_archivo_guias_remision+formato))

    limpiar_excel_openpyxl(os.path.join(ruta_archivos_base,nombre_archivo_entregas_ewm_cabezera+formato))
    limpiar_excel_openpyxl(os.path.join(ruta_archivos_base,nombre_archivo_entregas_ewm_cabezera+centros["centro"][0]+formato))
    limpiar_excel_openpyxl(os.path.join(ruta_archivos_base,nombre_archivo_entregas_ewm_cabezera+centros["centro"][1]+formato))
    limpiar_excel_openpyxl(os.path.join(ruta_archivos_base,nombre_archivo_entregas_ewm_cabezera+centros["centro"][2]+formato))

    limpiar_excel_openpyxl(os.path.join(ruta_archivos_base,nombre_archivo_entregas_ewm_cabezera_pendiente+centros_pendiente["centro"][0]+formato))
    limpiar_excel_openpyxl(os.path.join(ruta_archivos_base,nombre_archivo_entregas_ewm_cabezera_pendiente+centros_pendiente["centro"][1]+formato))
    limpiar_excel_openpyxl(os.path.join(ruta_archivos_base,nombre_archivo_entregas_ewm_cabezera_pendiente+centros_pendiente["centro"][2]+formato))
    limpiar_excel_openpyxl(os.path.join(ruta_archivos_base,nombre_archivo_entregas_ewm_cabezera_pendiente+centros_pendiente["centro"][3]+formato))

    limpiar_excel_openpyxl(os.path.join(ruta_archivos_base,nombre_archivo_entregas_ewm_detalle+formato))
    limpiar_excel_openpyxl(os.path.join(ruta_archivos_base,nombre_archivo_entregas_ewm_detalle+centros["centro"][0]+formato))
    limpiar_excel_openpyxl(os.path.join(ruta_archivos_base,nombre_archivo_entregas_ewm_detalle+centros["centro"][1]+formato))
    limpiar_excel_openpyxl(os.path.join(ruta_archivos_base,nombre_archivo_entregas_ewm_detalle+centros["centro"][2]+formato))

    limpiar_excel_openpyxl(os.path.join(ruta_archivos_base,nombre_archivo_entregas_ewm_detalle_pendiente+centros_pendiente["centro"][0]+formato))
    limpiar_excel_openpyxl(os.path.join(ruta_archivos_base,nombre_archivo_entregas_ewm_detalle_pendiente+centros_pendiente["centro"][1]+formato))
    limpiar_excel_openpyxl(os.path.join(ruta_archivos_base,nombre_archivo_entregas_ewm_detalle_pendiente+centros_pendiente["centro"][2]+formato))
    limpiar_excel_openpyxl(os.path.join(ruta_archivos_base,nombre_archivo_entregas_ewm_detalle_pendiente+centros_pendiente["centro"][3]+formato))

if __name__ == "__main__":
    main()

