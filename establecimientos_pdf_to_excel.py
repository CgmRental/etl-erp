import pdfplumber
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys
import re


def clean_text(text):
    """Limpia y normaliza el texto extraído."""
    if text is None:
        return ""
    # Eliminar espacios múltiples y saltos de línea extra
    text = re.sub(r'\s+', ' ', str(text))
    return text.strip()


def extract_establecimientos_from_pdf(pdf_path):
    """
    Extrae la información de establecimientos del PDF.
    
    Args:
        pdf_path: Ruta al archivo PDF
        
    Returns:
        DataFrame con los datos extraídos
    """
    all_rows = []
    
    print(f"Procesando: {pdf_path}")
    
    with pdfplumber.open(pdf_path) as pdf:
        print(f"Total de páginas: {len(pdf.pages)}")
        
        for page_num, page in enumerate(pdf.pages, 1):
            print(f"  Extrayendo página {page_num}/{len(pdf.pages)}...")
            
            # Extraer tablas de la página
            tables = page.extract_tables()
            
            for table in tables:
                if not table:
                    continue
                
                # Procesar cada fila de la tabla
                for row in table:
                    # Saltar filas vacías o headers
                    if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                        continue
                    
                    # Saltar headers de la tabla
                    if row[0] and 'Código' in str(row[0]):
                        continue
                    
                    # Saltar filas de paginación
                    if any(text in str(row) for text in ['Páginas:', 'Siguiente', 'a 30 de 100', 'SUNAT']):
                        continue
                    
                    # Verificar que tenga al menos 4 columnas
                    if len(row) >= 4:
                        codigo = clean_text(row[0])
                        tipo = clean_text(row[1])
                        direccion = clean_text(row[2])
                        actividad = clean_text(row[3]) if len(row) > 3 else ""
                        
                        # Solo agregar si tiene código
                        if codigo and codigo != '-':
                            all_rows.append({
                                'Código': codigo,
                                'Tipo de Establecimiento': tipo,
                                'Dirección': direccion,
                                'Actividad Económica': actividad
                            })
    
    # Crear DataFrame
    df = pd.DataFrame(all_rows)
    
    # Eliminar duplicados exactos
    df = df.drop_duplicates()
    
    print(f"\n✓ Registros extraídos: {len(df)}")
    
    return df


def save_to_excel_formatted(df, output_path, pdf_name):
    """
    Guarda el DataFrame en Excel con formato profesional.
    
    Args:
        df: DataFrame con los datos
        output_path: Ruta del archivo Excel de salida
        pdf_name: Nombre del PDF original
    """
    # Crear archivo Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Hoja principal con los datos
        df.to_excel(writer, sheet_name='Establecimientos', index=False)
        
        # Hoja de resumen
        resumen_data = {
            'Información': [
                'Archivo Origen',
                'Total de Establecimientos',
                'Fecha de Extracción',
                'Tipos de Establecimiento'
            ],
            'Valor': [
                pdf_name,
                len(df),
                pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S'),
                df['Tipo de Establecimiento'].nunique()
            ]
        }
        df_resumen = pd.DataFrame(resumen_data)
        df_resumen.to_excel(writer, sheet_name='Resumen', index=False)
        
        # Hoja de estadísticas por tipo
        if not df.empty and 'Tipo de Establecimiento' in df.columns:
            tipo_stats = df['Tipo de Establecimiento'].value_counts().reset_index()
            tipo_stats.columns = ['Tipo de Establecimiento', 'Cantidad']
            tipo_stats.to_excel(writer, sheet_name='Por Tipo', index=False)
    
    # Aplicar formato
    wb = load_workbook(output_path)
    
    # Definir estilos
    header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Formatear hoja "Establecimientos"
    if 'Establecimientos' in wb.sheetnames:
        ws = wb['Establecimientos']
        
        # Encabezados
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 12  # Código
        ws.column_dimensions['B'].width = 20  # Tipo
        ws.column_dimensions['C'].width = 80  # Dirección
        ws.column_dimensions['D'].width = 15  # Actividad
        
        # Aplicar bordes y alineación a todas las celdas con datos
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=4):
            for cell in row:
                cell.border = border
                if cell.column == 1:  # Código
                    cell.alignment = Alignment(horizontal='center', vertical='top')
                elif cell.column == 2:  # Tipo
                    cell.alignment = Alignment(horizontal='left', vertical='top')
                else:  # Dirección y Actividad
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Fijar primera fila
        ws.freeze_panes = 'A2'
        
        # Aplicar filtros
        ws.auto_filter.ref = ws.dimensions
    
    # Formatear hoja "Resumen"
    if 'Resumen' in wb.sheetnames:
        ws = wb['Resumen']
        
        for cell in ws[1]:
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40
        
        # Aplicar colores alternados
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            fill_color = 'F2F2F2' if row_idx % 2 == 0 else 'FFFFFF'
            for cell in row:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                cell.border = border
    
    # Formatear hoja "Por Tipo"
    if 'Por Tipo' in wb.sheetnames:
        ws = wb['Por Tipo']
        
        for cell in ws[1]:
            cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        
        # Formato a números
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            row[1].alignment = Alignment(horizontal='center')
            for cell in row:
                cell.border = border
    
    wb.save(output_path)
    print(f"✓ Archivo guardado con formato profesional")


def process_establecimientos_pdf(pdf_path, output_path=None):
    """
    Procesa el PDF de establecimientos y genera el Excel.
    
    Args:
        pdf_path: Ruta al archivo PDF
        output_path: Ruta del Excel de salida (opcional)
    """
    pdf_path = Path(pdf_path)
    
    if not pdf_path.exists():
        print(f"Error: El archivo '{pdf_path}' no existe.")
        return
    
    # Determinar ruta de salida
    if output_path is None:
        output_path = pdf_path.with_suffix('.xlsx')
    else:
        output_path = Path(output_path)
    
    print("=" * 70)
    print("EXTRACTOR DE ESTABLECIMIENTOS ANEXOS - PDF a Excel")
    print("=" * 70)
    print()
    
    try:
        # Extraer datos
        df = extract_establecimientos_from_pdf(pdf_path)
        
        if df.empty:
            print("\n⚠ No se encontraron datos para extraer.")
            return
        
        # Guardar en Excel
        print(f"\nGuardando en: {output_path}")
        save_to_excel_formatted(df, output_path, pdf_path.name)
        
        print()
        print("=" * 70)
        print("✓ PROCESO COMPLETADO EXITOSAMENTE")
        print("=" * 70)
        print(f"\nArchivo generado: {output_path}")
        print(f"Total de establecimientos: {len(df)}")
        print(f"Columnas extraídas: {', '.join(df.columns)}")
        
        # Mostrar muestra de datos
        if len(df) > 0:
            print("\nPrimeros 3 registros:")
            print(df.head(3).to_string(index=False))
        
    except Exception as e:
        print(f"\n✗ Error durante el procesamiento: {str(e)}")
        import traceback
        traceback.print_exc()


def main():
    """Función principal."""
    if len(sys.argv) < 2:
        print("=" * 70)
        print("EXTRACTOR DE ESTABLECIMIENTOS ANEXOS")
        print("=" * 70)
        print("\nUso:")
        print("  python establecimientos_pdf_to_excel.py <archivo.pdf> [salida.xlsx]")
        print("\nEjemplos:")
        print("  python establecimientos_pdf_to_excel.py Establecimientos_Anexos.pdf")
        print("  python establecimientos_pdf_to_excel.py input.pdf output.xlsx")
        print()
        return
    
    pdf_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    process_establecimientos_pdf(pdf_file, output_file)


if __name__ == '__main__':
    main()
